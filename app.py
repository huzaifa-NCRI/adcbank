from flask import Flask, request, send_file
from werkzeug.utils import secure_filename
from flask import Flask, request, send_file, render_template
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)

# Set the folder where uploaded files will be stored
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
DOWNLOAD_FOLDER = 'download_formatted_sheet'
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER

# Function to check if the file is allowed (for example, only Excel files)
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xls', 'xlsx'}

@app.route('/')
def index():
    return render_template('index.html')
    # return '''
    # <html>
    #     <body>
    #         <h2>Upload an Excel file</h2>
    #         <form method="POST" enctype="multipart/form-data" action="/upload">
    #             <input type="file" name="file">
    #             <input type="submit" value="Upload">
    #         </form>
    #     </body>
    # </html>
    # '''

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        error= "No file part"

    file = request.files['file']

    if file.filename == '':
        error= "No selected file"

    if file and allowed_file(file.filename):
        # Securely save the uploaded file
        source_file = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], source_file)
        output_file = os.path.splitext(source_file)[0] + '_formatted.xlsx'
        download_file_path = os.path.join(app.config['DOWNLOAD_FOLDER'], output_file)

        if file_path:
            file.save(file_path)
            source_workbook = openpyxl.load_workbook(file_path)
        else:
            # Handle the case where file.save did not return a valid file path
            print("File was not saved correctly.")
        
        # file_path = file.save(os.path.join(app.config['UPLOAD_FOLDER'], source_file))

        # source_workbook = openpyxl.load_workbook(file_path)
        source_sheet = source_workbook.active
        # Create a new Excel file to save the transformed data
        base_filename, file_extension = os.path.splitext(os.path.basename(source_file))
        # Create a new Excel file with the processed name
        output_file = f"{base_filename}_formatted.xlsx"
        # output_file = "output.xlsx"
        output_workbook = openpyxl.Workbook()   
        output_sheet = output_workbook.active


        col1_values = []
        col3_values = []
        col5_values = []

        # Iterate through the first 5 rows of the input sheet
        for row_index, row in enumerate(source_workbook.active.iter_rows(max_row=11), start=1):
            # Extract values from columns 1, 3, and 5
            col1_value = row[0].value
            col3_value = row[2].value
            col5_value = row[4].value

            # Append the values to their respective lists
            col1_values.append(col1_value)
            col3_values.append(col3_value)
            col5_values.append(col5_value)

        # Append all the transposed values to the output sheet in the desired order
        transposed_vals = col1_values + col3_values + col5_values

        def copy_and_transform_data(start_row, end_row, start_column, end_column, output_sheet):
            for row in source_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
                output_row = []
                for cell in row:
                    output_row.append(cell.value)
                output_sheet.append(output_row)


        # Define the column ranges to process
        column_ranges = [(2, 2), (4, 4), (6, 6)]  # Adjust as needed
        # column_header = [(1, 1), (3, 3), (5, 5)]

        # Create a new Workbook for the transposed data
        transposed_workbook = openpyxl.Workbook()
        transposed_sheet = transposed_workbook.active

        # Define the number of rows to process in each iteration
        chunk_size = 11

        # Define the number of rows to skip after each chunk
        skip_rows = 6

        # Initialize start_row
        start_row = 1
        # Iterate over the entire sheet in chunks of 9 rows at a time
        while start_row <= source_sheet.max_row:
            end_row = min(start_row + chunk_size - 1, source_sheet.max_row)
            end_head = 1
            # for start, end in column_header:
            #     copy_and_transform_header(start_row, end_head, start, end, transposed_sheet)

            for start, end in column_ranges:
                copy_and_transform_data(start_row, end_row, start, end, transposed_sheet)

            # Update start_row to skip the next 3 rows
            start_row = end_row + skip_rows + 1

        # Transpose the data
        transposed_data = list(map(list, zip(*transposed_sheet.iter_rows(values_only=True))))

        # Create a new Workbook for the transposed data
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active

        # Write the transposed data to the output sheet
        columns_per_row = 33
        start_row = source_sheet.min_row + 14
        last_row = source_sheet.min_row + 14
        start_header = source_sheet.min_row + 13
        last_header = source_sheet.min_row + 13

        counter = 0
        light_blue_fill = PatternFill(start_color="ADD8E6", end_color="00CCFF", fill_type="solid")
        italic_font = Font(italic=True)
        # combined_head = transposed_vals
        # output_sheet.append(combined_head)
        for row in transposed_data:
            
            while len(row) > 0:
                
                # Append the current row with 26 columns per row to the output sheet

                
                combined_head = transposed_vals
                combined_row = row[:columns_per_row]  # Take the first part of 'row'
                row = row[columns_per_row:]  # Remove the first part from 'row'
                
                # Prepare the 'output_row' from 'source_sheet.iter_rows'
                for source_row in source_sheet.iter_rows(min_row=start_row, max_row=last_row):
                    output_row = []
                    for cell in source_row:
                        output_row.append(cell.value)

                # Prepare the 'output_row' from 'source_sheet.iter_rows'
                for source_row in source_sheet.iter_rows(min_row=start_header, max_row=last_header):
                    output_head = []
                    for cell in source_row:
                        output_head.append(cell.value)
                
                # Concatenate 'combined_row' and 'output_row'
                combined_row.extend(output_row)
                combined_head.extend(output_head)

                start_row = start_row + 17
                last_row = last_row + 17
                # start_header = start_header + 14
                # last_header = last_header + 14
                
                # Append the combined row to the output sheet
                if counter == 0:
                    # remove space from CIF
                    removeSpace_col0 = combined_head[0].strip()
                    combined_head[0] = removeSpace_col0
                    # Split one col into 2
                    split_col11 = combined_head[11].split("/")
                    combined_head[11:12] = split_col11
                    split_col13 = combined_head[13].split("/")
                    combined_head[13:14] = split_col13
                    split_col28 = combined_head[28].split("/")
                    combined_head[28:29] = split_col28

                    # remove space from Passport and Account_no
                    removeSpace_col12 = combined_head[12].strip()
                    combined_head[12] = removeSpace_col12
                    removeSpace_col36 = combined_head[36].strip()
                    combined_head[36] = removeSpace_col36


                    output_sheet.append(combined_head)
                    counter += 1
                    # Apply the blue fill to the header row cells
                    for cell in output_sheet[1]:
                        cell.fill = light_blue_fill 
                        cell.font = italic_font

                # remove space from CIF
                removeSpace_col0 = combined_row[0].strip()
                combined_row[0] = removeSpace_col0
                # Split columns col11, col13, and col28 using "/"
                split_col11 = combined_row[11].split("/")
                combined_row[11:12] = split_col11
                split_col13 = combined_row[13].split("/")
                combined_row[13:14] = split_col13
                split_col28 = combined_row[28].split("/")
                combined_row[28:29] = split_col28

                # remove space from Passport and Account_no
                removeSpace_col12 = combined_row[12].strip()
                combined_row[12] = removeSpace_col12
                removeSpace_col36 = combined_row[36].strip()
                combined_row[36] = removeSpace_col36
                

                output_sheet.append(combined_row)
        output_workbook.save(download_file_path)
        


        return send_file(download_file_path, as_attachment=True)
        # return f"File '{output_file}' uploaded and processed successfully. <a href='/download/{output_file}'>Download</a>"

    else:
        # return "File type not allowed. Please upload an Excel file."
        return render_template('upload.html', error=error)
  

@app.route('/download/<filename>')
def download_file(filename):
    download_path = os.path.join(app.config['DOWNLOAD_FOLDER'], filename)
    if os.path.exists(download_path):
        return send_file(download_path, as_attachment=True)
    else:
        return "File not found"



if __name__ == '__main__':
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    if not os.path.exists(DOWNLOAD_FOLDER):
        os.makedirs(DOWNLOAD_FOLDER)
    app.run(debug=True)
